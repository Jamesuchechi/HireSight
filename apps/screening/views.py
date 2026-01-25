"""
Views for screening system with Mistral AI integration.
"""
import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.views import View
from django.views.generic import (
    CreateView, ListView, DetailView, UpdateView, FormView, TemplateView
)
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.db.models import Q, Avg, Count, Prefetch
from django.db import transaction
from django.utils import timezone
from django.core.exceptions import PermissionDenied
from django.core.files.storage import default_storage
from django import forms
from django.core.validators import FileExtensionValidator

from .models import (
    ScreeningSession, ScreeningResult, ScreeningCriteria,
    ScreeningStatus, ScreeningResultStatus, PipelineIntegration, PipelineStatus,
    ProgressUpdate, ProgressUpdateType, AIInsight, InsightType, InsightFeedback
)
from .forms import (
    ScreeningSessionForm, ScreeningCriteriaForm, BulkResumeUploadForm,
    ScreeningResultFilterForm, ScreeningResultExportForm,
    ScreeningSessionUpdateForm, ScreeningResultNoteForm,
    PushToPipelineForm, BulkPushToPipelineForm
)
from apps.resumes.models import Resume
from apps.jobs.models import Job
from apps.accounts.models import CompanyProfile
from apps.applications.models import Application, ApplicationStatus
from .services import ApplicationDataService
from apps.screening.tasks import process_resume_screening


class JobApplicationScreeningForm(forms.Form):
    """Form backing the job application screening UI."""

    title = forms.CharField(
        max_length=255,
        label="Screening Session Title",
        widget=forms.TextInput(attrs={'class': 'form-input w-full'})
    )
    required_skills = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-input w-full'}),
        help_text="Comma-separated required skills"
    )
    nice_to_have_skills = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-input w-full'}),
        help_text="Comma-separated nice-to-have skills"
    )
    min_experience_years = forms.FloatField(
        required=False,
        min_value=0,
        widget=forms.NumberInput(attrs={'class': 'form-input w-full', 'step': 0.5})
    )
    max_experience_years = forms.FloatField(
        required=False,
        min_value=0,
        widget=forms.NumberInput(attrs={'class': 'form-input w-full', 'step': 0.5})
    )
    weight_skills = forms.FloatField(
        min_value=0,
        max_value=1,
        initial=0.3,
        required=False,
        widget=forms.NumberInput(attrs={'class': 'form-input w-full', 'step': 0.05})
    )
    weight_experience = forms.FloatField(
        min_value=0,
        max_value=1,
        initial=0.2,
        required=False,
        widget=forms.NumberInput(attrs={'class': 'form-input w-full', 'step': 0.05})
    )
    weight_education = forms.FloatField(
        min_value=0,
        max_value=1,
        initial=0.2,
        required=False,
        widget=forms.NumberInput(attrs={'class': 'form-input w-full', 'step': 0.05})
    )
    weight_keywords = forms.FloatField(
        min_value=0,
        max_value=1,
        initial=0.1,
        required=False,
        widget=forms.NumberInput(attrs={'class': 'form-input w-full', 'step': 0.05})
    )
    weight_screening_questions = forms.FloatField(
        min_value=0,
        max_value=1,
        initial=0.1,
        required=False,
        widget=forms.NumberInput(attrs={'class': 'form-input w-full', 'step': 0.05})
    )
    weight_assessments = forms.FloatField(
        min_value=0,
        max_value=1,
        initial=0.1,
        required=False,
        widget=forms.NumberInput(attrs={'class': 'form-input w-full', 'step': 0.05})
    )
    custom_keywords = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-input w-full'}),
        help_text="Comma-separated keywords to boost"
    )
    screening_questions_config = forms.JSONField(
        required=False,
        widget=forms.Textarea(attrs={'class': 'form-textarea w-full', 'rows': 3}),
        help_text="Optional screening question configuration (JSON)"
    )

    def clean_required_skills(self):
        return [skill.strip() for skill in (self.cleaned_data.get('required_skills') or "").split(',') if skill.strip()]

    def clean_nice_to_have_skills(self):
        return [skill.strip() for skill in (self.cleaned_data.get('nice_to_have_skills') or "").split(',') if skill.strip()]

    def clean_custom_keywords(self):
        return [keyword.strip() for keyword in (self.cleaned_data.get('custom_keywords') or "").split(',') if keyword.strip()]


logger = logging.getLogger(__name__)


def _resolve_nested_value(source, path, default=0):
    """Safely resolve nested dict/object attributes for match details."""
    current = source
    for key in path.split('.'):
        if isinstance(current, dict):
            current = current.get(key)
        else:
            current = getattr(current, key, None)
        if current is None:
            return default
    return current if current is not None else default


def _safe_percent(value, already_percent=False):
    """Clamp numeric values to a 0-100 percentage."""
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return 0.0
    percent = numeric if already_percent else numeric * 100
    return max(0.0, min(percent, 100.0))


def _criteria_weight(criteria, field, default):
    """Safely read a criteria weight (0-1)."""
    if not criteria:
        return max(0.0, min(default, 1.0))
    value = getattr(criteria, field, default)
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        numeric = default
    return max(0.0, min(numeric, 1.0))


def _normalize_screening_answers(answer_list):
    """Ensure screening answers expose a consistent question/answer display."""
    normalized = []
    for item in answer_list or []:
        if not isinstance(item, dict):
            continue
        clean = item.copy()
        clean['display_question'] = (
            clean.get('question')
            or clean.get('question_text')
            or clean.get('label')
            or "Question"
        )
        clean['display_answer'] = (
            clean.get('answer')
            or clean.get('response')
            or clean.get('value')
        )
        normalized.append(clean)
    return normalized


def _analysis_percent(details, key):
    """Extract a percentage from a stored analysis block."""
    if not isinstance(details, dict):
        return 0.0
    block = details.get(key)
    if isinstance(block, dict):
        # Check for nested 'score' field in screening_answers_analysis and assessments_analysis
        if key in ['screening_answers_analysis', 'assessments_analysis']:
            value = block.get('score')
        else:
            value = block.get('overall_score')
            if value is None:
                value = block.get('score')
        if value is None:
            value = block.get('value')
    else:
        value = block
    if value is None:
        return 0.0
    return _safe_percent(value, already_percent=True)


def queue_screening_for_session(session, requeue_pending=False, applications=None):
    """Create screening results for job applications and queue processing tasks."""
    response = {
        'created': 0,
        'queued_new': 0,
        'requeued': 0,
        'info': None,
        'error': None
    }

    if session.job is None:
        response['error'] = "Session must be tied to a job before screening candidates."
        return response

    if applications is None:
        applications = Application.objects.filter(job=session.job).select_related('resume').order_by('-applied_at')

    if not applications.exists():
        response['info'] = "No applications exist for this job yet."
        return response

    existing_application_ids = set(session.results.values_list('application_id', flat=True))
    new_result_ids = []
    for application in applications:
        if application.id in existing_application_ids:
            continue

        file_path = ''
        if application.resume and getattr(application.resume, 'file', None):
            file_path = application.resume.file.name

        result = ScreeningResult.objects.create(
            session=session,
            application=application,
            resume=application.resume,
            job=session.job,
            file_path=file_path,
            status=ScreeningResultStatus.PENDING
        )
        response['created'] += 1
        new_result_ids.append(result.id)

        if file_path:
            process_resume_screening.delay(result.id)
            response['queued_new'] += 1
        else:
            logger.warning("Application %s has no resume file; screening task skipped", application.id)

    statuses_to_requeue = [ScreeningResultStatus.FAILED]
    if requeue_pending:
        statuses_to_requeue.append(ScreeningResultStatus.PENDING)

    pending_results = session.results.filter(status__in=statuses_to_requeue)
    if new_result_ids:
        pending_results = pending_results.exclude(id__in=new_result_ids)

    for pending in pending_results:
        if not pending.resume and not pending.file_path:
            continue

        process_resume_screening.delay(pending.id)
        response['requeued'] += 1

    total_results = session.results.count()
    session.total_resumes = total_results
    session.save(update_fields=['total_resumes'])

    if response['queued_new'] or response['requeued']:
        session.start_processing()
        response['info'] = None
    elif response['created'] and not response['queued_new']:
        response['info'] = "Results created but no resume files were available to process yet."
    elif not response['created']:
        response['info'] = "All applicants have already been queued for screening."

    return response


class CompanyOnlyMixin:
    """Mixin to ensure only company users can access screening."""

    def dispatch(self, request, *args, **kwargs):
        if request.user.account_type != 'company':
            messages.error(request, "Only company accounts can access screening features.")
            return redirect('dashboard:dashboard_home')
        return super().dispatch(request, *args, **kwargs)


class ScreeningOwnerMixin:
    """Mixin to ensure user owns the screening session."""

    def get_queryset(self):
        queryset = super().get_queryset()
        # Optimize queries
        queryset = queryset.select_related(
            'company',
            'company__user',
            'job',
            'created_by'
        ).prefetch_related(
            Prefetch('results', queryset=ScreeningResult.objects.select_related('resume', 'job')),
            'criteria'
        )
        return queryset.filter(company__user=self.request.user)

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if obj.company.user != self.request.user:
            logger.warning(f"Unauthorized access attempt to session {obj.id} by user {self.request.user.id}")
            raise PermissionDenied("You can only access your own screening sessions.")
        return obj


# ===========================
# Screening Session Views
# ===========================

class ScreeningSessionListView(LoginRequiredMixin, CompanyOnlyMixin, ScreeningOwnerMixin, ListView):
    """List all screening sessions for the company."""
    
    model = ScreeningSession
    template_name = 'screening/session_list.html'
    context_object_name = 'sessions'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(job__title__icontains=search)
            )
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        """Provide job-specific applications, summary stats, and current filter state."""
        context = super().get_context_data(**kwargs)
        
        # Add statistics
        queryset = self.get_queryset()
        context['stats'] = {
            'total': queryset.count(),
            'in_progress': queryset.filter(status__in=[
                ScreeningStatus.PENDING,
                ScreeningStatus.PROCESSING
            ]).count(),
            'completed': queryset.filter(status=ScreeningStatus.COMPLETED).count(),
            'failed': queryset.filter(status=ScreeningStatus.FAILED).count(),
        }
        
        return context


class ScreeningSessionCreateView(LoginRequiredMixin, CompanyOnlyMixin, CreateView):
    """Create a new screening session with criteria."""
    
    model = ScreeningSession
    form_class = ScreeningSessionForm
    template_name = 'screening/session_create.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        try:
            kwargs['company'] = self.request.user.company_profile
        except CompanyProfile.DoesNotExist:
            messages.error(self.request, "Your company profile is not set up. Please complete your profile first.")
            raise PermissionDenied("Company profile required")
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add available jobs to context for display
        try:
            company = self.request.user.company_profile
            context['available_jobs'] = Job.objects.filter(
                company=company,
                status='active'
            ).order_by('-created_at')
        except CompanyProfile.DoesNotExist:
            context['available_jobs'] = []
        return context

    @transaction.atomic
    def form_valid(self, form):
        """Handle valid form submission."""
        try:
            # Save session
            session = form.save(commit=False)
            try:
                session.company = self.request.user.company_profile
            except CompanyProfile.DoesNotExist:
                messages.error(self.request, "Your company profile is not set up. Please complete your profile first.")
                return self.form_invalid(form)
            
            session.created_by = self.request.user
            session.status = ScreeningStatus.PENDING
            session.save()
            
            # Create empty criteria for this session
            # This ensures the criteria exists when we redirect
            ScreeningCriteria.objects.create(
                session=session,
                required_skills=[],
                nice_to_have_skills=[],
                min_experience_years=0,
                required_education=[],
                custom_keywords=[],
                weight_skills=0.3,
                weight_experience=0.2,
                weight_education=0.2,
                weight_keywords=0.1,
                weight_screening_questions=0.1,
                weight_assessments=0.1
            )
            
            logger.info(f"Screening session created: {session.id} by user {self.request.user.id}")
            messages.success(self.request, "Screening session created! Now define screening criteria.")
            
            # Store session ID in session for next step
            self.request.session['current_screening_session'] = str(session.id)
            
            # Redirect to criteria setup
            return redirect('screening:criteria_setup', session_id=session.pk)
            
        except Exception as e:
            logger.error(f"Error creating screening session: {str(e)}", exc_info=True)
            messages.error(self.request, f"An error occurred while creating the session: {str(e)}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Handle invalid form submission."""
        logger.error(f"Form validation errors: {form.errors}")
        logger.error(f"Form data: {form.data}")
        messages.error(self.request, "Please correct the errors below.")
        return super().form_invalid(form)


class JobApplicationScreeningView(LoginRequiredMixin, CompanyOnlyMixin, FormView):
    """View that lets a recruiter select a job's applications and trigger screening."""

    template_name = 'screening/job_application_screening.html'
    form_class = JobApplicationScreeningForm

    def dispatch(self, request, *args, **kwargs):
        self.job = get_object_or_404(
            Job,
            pk=self.kwargs.get('job_id'),
            company__user=request.user
        )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        applications_data = self._assemble_applications()
        filtered = self._apply_filters(applications_data)

        context.update({
            'job': self.job,
            'applications_data': filtered,
            'summary': self._build_summary(applications_data),
            'active_filters': self._get_filter_params(),
            'status_choices': ApplicationStatus.choices
        })
        return context

    def _assemble_applications(self):
        """Build the context rows by enriching each application."""
        rows = []
        queryset = ApplicationDataService.get_job_applications(self.job.id)
        for application in queryset:
            screening_data = ApplicationDataService.get_application_screening_data(application)
            rows.append({
                'application': application,
                'screening_data': screening_data,
                'has_resume': bool(application.resume),
                'answers_completed': bool(screening_data.get('screening_answers')),
                'assessments_count': len(screening_data.get('assessment_results', [])),
            })
        return rows

    def _get_filter_params(self):
        return {
            'has_resume': self.request.GET.get('has_resume'),
            'screening_answers': self.request.GET.get('screening_answers'),
            'assessments': self.request.GET.get('assessments'),
            'status': self.request.GET.get('status')
        }

    def _apply_filters(self, rows):
        """Filter the assembled applications based on the active query parameters."""
        params = self._get_filter_params()
        filtered = rows

        if params['has_resume'] == '1':
            filtered = [row for row in filtered if row['has_resume']]
        if params['screening_answers'] == 'complete':
            filtered = [row for row in filtered if row['answers_completed']]
        elif params['screening_answers'] == 'incomplete':
            filtered = [row for row in filtered if not row['answers_completed']]
        if params['assessments'] == 'taken':
            filtered = [row for row in filtered if row['assessments_count'] > 0]
        elif params['assessments'] == 'none':
            filtered = [row for row in filtered if row['assessments_count'] == 0]
        if params['status']:
            filtered = [row for row in filtered if row['application'].status == params['status']]

        return filtered

    def _build_summary(self, rows):
        total = len(rows)
        ready = sum(1 for row in rows if row['has_resume'] and row['answers_completed'])
        pending = total - ready
        return {
            'total_applicants': total,
            'ready_to_screen': ready,
            'pending_data': pending,
            'screened_count': total - pending
        }

    def form_valid(self, form):
        selected_ids = self.request.POST.getlist('selected_applications')
        if not selected_ids:
            form.add_error(None, "Select at least one application to screen.")
            return self.form_invalid(form)

        applications = list(Application.objects.filter(
            id__in=selected_ids,
            job=self.job
        ).select_related('resume'))

        if not applications:
            form.add_error(None, "No valid applications selected.")
            return self.form_invalid(form)

        session = ScreeningSession.objects.create(
            title=form.cleaned_data['title'],
            company=self.job.company,
            job=self.job,
            created_by=self.request.user,
            status=ScreeningStatus.PENDING
        )

        criteria = ScreeningCriteria.objects.create(
            session=session,
            required_skills=form.cleaned_data['required_skills'],
            nice_to_have_skills=form.cleaned_data['nice_to_have_skills'],
            min_experience_years=form.cleaned_data.get('min_experience_years') or 0,
            max_experience_years=form.cleaned_data.get('max_experience_years'),
            custom_keywords=form.cleaned_data.get('custom_keywords') or [],
            weight_skills=form.cleaned_data.get('weight_skills') or 0.3,
            weight_experience=form.cleaned_data.get('weight_experience') or 0.2,
            weight_education=form.cleaned_data.get('weight_education') or 0.2,
            weight_keywords=form.cleaned_data.get('weight_keywords') or 0.1,
            weight_screening_questions=form.cleaned_data.get('weight_screening_questions') or 0.1,
            weight_assessments=form.cleaned_data.get('weight_assessments') or 0.1,
            screening_questions_config=form.cleaned_data.get('screening_questions_config') or {}
        )

        results_created = 0
        for application in applications:
            file_path = ''
            if application.resume and application.resume.file:
                file_path = application.resume.file.name

            result = ScreeningResult.objects.create(
                session=session,
                application=application,
                resume=application.resume,
                job=self.job,
                file_path=file_path,
                status=ScreeningResultStatus.PENDING
            )
            results_created += 1

            if file_path:
                process_resume_screening.delay(result.id)
            else:
                logger.warning("No resume file for application %s; skipping background task", application.id)

        session.total_resumes = results_created
        session.status = ScreeningStatus.PROCESSING if results_created else ScreeningStatus.PENDING
        session.save(update_fields=['total_resumes', 'status'])

        messages.success(self.request, f"{results_created} candidate(s) queued for screening.")
        return redirect('screening:session_detail', session_id=session.pk)


class ScreenSingleApplicationView(LoginRequiredMixin, CompanyOnlyMixin, TemplateView):
    """Render screening details for a single application and allow manual review."""

    template_name = 'screening/screen_single_application.html'

    def dispatch(self, request, *args, **kwargs):
        self.application = get_object_or_404(
            Application.objects.select_related('job__company', 'applicant__personal_profile'),
            id=self.kwargs.get('application_id')
        )
        if self.application.job.company.user_id != request.user.id:
            raise PermissionDenied("You do not own the job tied to this application.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['application'] = self.application
        context['screening_data'] = ApplicationDataService.get_application_screening_data(self.application)
        return context


class ScreeningCriteriaSetupView(LoginRequiredMixin, CompanyOnlyMixin, FormView):
    """Set up screening criteria for a session."""
    
    form_class = ScreeningCriteriaForm
    template_name = 'screening/criteria_setup.html'

    def get_session(self):
        """Get the screening session."""
        session_id = self.kwargs.get('session_id')
        try:
            session = ScreeningSession.objects.select_related('company').get(
                pk=session_id,
                company__user=self.request.user
            )
            return session
        except ScreeningSession.DoesNotExist:
            logger.error(f"Session {session_id} not found for user {self.request.user.id}")
            messages.error(self.request, "Screening session not found.")
            return None

    def dispatch(self, request, *args, **kwargs):
        """Check if session exists before processing."""
        session = self.get_session()
        if session is None:
            return redirect('screening:session_list')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        session = self.get_session()
        kwargs['session'] = session
        
        # Get existing criteria if it exists
        try:
            criteria = ScreeningCriteria.objects.get(session=session)
            kwargs['instance'] = criteria
        except ScreeningCriteria.DoesNotExist:
            # Create new criteria if it doesn't exist
            pass
        
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['session'] = self.get_session()
        
        # Check if criteria already exists
        try:
            context['criteria'] = ScreeningCriteria.objects.get(session=context['session'])
        except ScreeningCriteria.DoesNotExist:
            context['criteria'] = None
        
        return context

    @transaction.atomic
    def form_valid(self, form):
        """Save criteria and redirect to upload."""
        try:
            session = self.get_session()
            
            # Check if criteria already exists
            try:
                criteria = ScreeningCriteria.objects.get(session=session)
                # Update existing criteria
                for field, value in form.cleaned_data.items():
                    setattr(criteria, field, value)
                criteria.save()
                logger.info(f"Updated criteria for session {session.id}")
            except ScreeningCriteria.DoesNotExist:
                # Create new criteria
                criteria = form.save(commit=False)
                criteria.session = session
                criteria.save()
                logger.info(f"Created new criteria for session {session.id}")
            
            messages.success(self.request, "Screening criteria saved! Now upload resumes.")
            
            # Clear session data
            if 'current_screening_session' in self.request.session:
                del self.request.session['current_screening_session']
            
            return redirect('screening:supplement_resumes', session_id=session.pk)
            
        except Exception as e:
            logger.error(f"Error saving criteria: {str(e)}", exc_info=True)
            messages.error(self.request, f"An error occurred while saving criteria: {str(e)}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Handle invalid form."""
        logger.error(f"Criteria form errors: {form.errors}")
        messages.error(self.request, "Please correct the errors in the form.")
        return super().form_invalid(form)


class SupplementResumeForm(forms.Form):
    email = forms.EmailField(
        label="Applicant email",
        widget=forms.EmailInput(attrs={'class': 'form-input w-full'}),
        help_text="Email used during application"
    )
    resume = forms.FileField(
        label="Resume file",
        widget=forms.FileInput(attrs={'class': 'form-input w-full'}),
        help_text="Upload resumes only for applicants who applied without one",
        validators=[FileExtensionValidator(allowed_extensions=['pdf', 'doc', 'docx'])]
    )


class SupplementMissingResumesView(LoginRequiredMixin, CompanyOnlyMixin, FormView):
    """Supplement missing resumes for a screening session."""
    
    form_class = SupplementResumeForm
    template_name = 'screening/supplement_missing_resumes.html'

    def get_session(self):
        session_id = self.kwargs.get('session_id')
        return get_object_or_404(
            ScreeningSession,
            pk=session_id,
            company__user=self.request.user
        )

    def dispatch(self, request, *args, **kwargs):
        """Only show the supplement page when missing resumes exist."""
        session = self.get_session()
        self.missing_applications = Application.objects.filter(
            job=session.job,
            resume__isnull=True
        ).select_related('applicant__personal_profile').order_by('-applied_at')

        if self.missing_applications.count() == 0:
            queue_result = queue_screening_for_session(session)
            if queue_result['error']:
                messages.error(request, queue_result['error'])
            elif queue_result['queued_new'] or queue_result['requeued']:
                message_parts = []
                if queue_result['queued_new']:
                    message_parts.append(f"{queue_result['queued_new']} resume(s) queued for screening.")
                if queue_result['requeued']:
                    message_parts.append(f"{queue_result['requeued']} previous result(s) requeued.")
                messages.success(request, " ".join(message_parts))
            else:
                messages.info(request, queue_result.get('info') or "No resumes were queued for screening.")
            return redirect('screening:session_detail', session_id=session.pk)

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session = self.get_session()
        missing_apps = getattr(self, 'missing_applications', None)
        if missing_apps is None:
            missing_apps = Application.objects.filter(
                job=session.job,
                resume__isnull=True
            ).select_related('applicant__personal_profile').order_by('-applied_at')

        context.update({
            'session': session,
            'missing_applications': missing_apps,
            'missing_count': missing_apps.count()
        })
        return context

    def post(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return self.handle_ajax_upload()
        return super().post(request, *args, **kwargs)

    @transaction.atomic
    def handle_ajax_upload(self):
        session = self.get_session()
        email = self.request.POST.get('email', '').strip()
        uploaded_file = self.request.FILES.get('resume')

        if not uploaded_file or not email:
            return JsonResponse({
                'success': False,
                'message': 'Email and resume file are required.'
            }, status=400)

        return self._process_resume_for_application(session, email, uploaded_file, ajax=True)

    def form_valid(self, form):
        session = self.get_session()
        email = form.cleaned_data['email']
        uploaded_file = form.cleaned_data['resume']

        result = self._process_resume_for_application(session, email, uploaded_file)
        if isinstance(result, JsonResponse):
            form.add_error(None, result.json().get('message'))
            return self.form_invalid(form)

        messages.success(self.request, result['message'])
        return redirect('screening:supplement_resumes', session_id=session.pk)

    def _process_resume_for_application(self, session, email, uploaded_file, ajax=False):
        application = Application.objects.filter(
            job=session.job,
            applicant__email__iexact=email
        ).select_related('applicant').first()

        if not application:
            payload = {
                'success': False,
                'message': 'No application found for this candidate.'
            }
            if ajax:
                return JsonResponse(payload, status=404)
            return payload

        resume = Resume.objects.create(
            user=application.applicant,
            title=f"Supplemented Resume ({application.job.title})",
            original_filename=uploaded_file.name,
            status='uploaded',
            file_size=uploaded_file.size
        )
        resume.file.save(uploaded_file.name, uploaded_file, save=True)

        application.resume = resume
        tags = application.tags or []
        if 'resume_supplemented' not in tags:
            tags.append('resume_supplemented')
        application.tags = tags
        application.last_activity_at = timezone.now()
        application.save(update_fields=['resume', 'tags', 'last_activity_at'])

        # Ensure there is a screening result available for this application
        result = ScreeningResult.objects.filter(session=session, application=application).first()
        file_path = resume.file.name if resume.file else ''
        if result:
            result.resume = resume
            result.file_path = file_path
            result.error_message = ''
            result.status = ScreeningResultStatus.PENDING
            result.save(update_fields=['resume', 'file_path', 'status', 'error_message'])
        else:
            result = ScreeningResult.objects.create(
                session=session,
                application=application,
                resume=resume,
                job=session.job,
                file_path=file_path,
                status=ScreeningResultStatus.PENDING
            )
            session.total_resumes = session.results.count()
            session.save(update_fields=['total_resumes'])

        # Kick off processing immediately
        process_resume_screening.delay(result.id)
        session.start_processing()

        message = 'Resume linked to application successfully.'
        if ajax:
            return JsonResponse({'success': True, 'message': message})
        return {'success': True, 'message': message}


class ScreeningSessionDetailView(LoginRequiredMixin, CompanyOnlyMixin, ScreeningOwnerMixin, DetailView):
    """View detailed screening session with progress."""
    
    model = ScreeningSession
    template_name = 'screening/session_detail.html'
    context_object_name = 'session'
    pk_url_kwarg = 'session_id'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session = self.object
        
        # Get results with optimization
        results = session.results.select_related('resume', 'job').order_by('-match_score')
        
        context['results'] = results[:10]  # Top 10 for preview
        context['total_results'] = results.count()
        
        # Statistics
        context['stats'] = {
            'total': session.total_resumes,
            'processed': session.processed_resumes,
            'failed': session.failed_resumes,
            'pending': session.total_resumes - session.processed_resumes - session.failed_resumes,
            'average_score': session.average_match_score or 0,
            'high_matches': results.filter(match_score__gte=80).count(),
            'shortlisted': results.filter(is_shortlisted=True).count(),
        }
        
        # Progress percentage
        if session.total_resumes > 0:
            context['progress'] = (session.processed_resumes / session.total_resumes) * 100
        else:
            context['progress'] = 0
        
        # Get criteria
        try:
            context['criteria'] = session.criteria
        except ScreeningCriteria.DoesNotExist:
            context['criteria'] = None


        if session.job:
            job_apps = Application.objects.filter(job=session.job)
            context['missing_applications_count'] = job_apps.filter(resume__isnull=True).count()
            context['ready_applications_count'] = job_apps.filter(resume__isnull=False).count()
        else:
            context['missing_applications_count'] = 0
            context['ready_applications_count'] = 0
        
        return context


class ScreeningSessionStartProcessingView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Trigger processing for all ready resumes belonging to a session."""

    def post(self, request, *args, **kwargs):
        session = get_object_or_404(
            ScreeningSession,
            pk=self.kwargs.get('session_id'),
            company__user=request.user
        )

        if not getattr(session, 'criteria', None):
            messages.error(request, "Set up screening criteria before starting processing.")
            return redirect('screening:criteria_setup', session_id=session.pk)

        reprocess_all = request.POST.get('reprocess_all') == '1'
        if reprocess_all:
            session.results.filter(status=ScreeningResultStatus.COMPLETED).update(
                status=ScreeningResultStatus.PENDING,
                match_score=0,
                match_details={},
                screening_answers=[],
                assessment_data=[],
                error_message='',
                processed_at=None
            )
            session.processed_resumes = 0
            session.failed_resumes = 0
            session.average_match_score = None
            session.status = ScreeningStatus.PENDING
            session.save(update_fields=['processed_resumes', 'failed_resumes', 'average_match_score', 'status'])
            messages.info(request, "All completed results have been reset and will be reprocessed.")

        queue_result = queue_screening_for_session(session, requeue_pending=True)
        if queue_result['error']:
            messages.error(request, queue_result['error'])
        elif queue_result['queued_new'] or queue_result['requeued']:
            message_parts = []
            if queue_result['queued_new']:
                message_parts.append(f"{queue_result['queued_new']} resume(s) queued for screening.")
            if queue_result['requeued']:
                message_parts.append(f"{queue_result['requeued']} previous result(s) requeued.")
            messages.success(request, " ".join(message_parts))
        else:
            messages.info(request, queue_result.get('info') or "No resumes were queued for screening.")

        return redirect('screening:session_detail', session_id=session.pk)


class ScreeningResultsView(LoginRequiredMixin, CompanyOnlyMixin, ListView):
    """View all screening results for a session."""
    
    model = ScreeningResult
    template_name = 'screening/results_list.html'
    context_object_name = 'results'
    paginate_by = 20

    def get_session(self):
        """Get the screening session."""
        session_id = self.kwargs.get('session_id')
        return get_object_or_404(
            ScreeningSession,
            pk=session_id,
            company__user=self.request.user
        )

    def get_queryset(self):
        session = self.get_session()
        queryset = ScreeningResult.objects.filter(session=session).select_related(
            'resume',
            'resume__user',
            'resume__user__personal_profile',
            'job'
        )
        
        # Apply filters
        filter_form = ScreeningResultFilterForm(self.request.GET)
        if filter_form.is_valid():
            # Score range
            score_range = filter_form.cleaned_data.get('score_range')
            if score_range:
                min_score = filter_form.cleaned_data.get('score_range_min')
                max_score = filter_form.cleaned_data.get('score_range_max')
                if min_score is not None and max_score is not None:
                    queryset = queryset.filter(
                        match_score__gte=min_score,
                        match_score__lte=max_score
                    )
            
            # Custom min/max
            min_score = filter_form.cleaned_data.get('min_score')
            max_score = filter_form.cleaned_data.get('max_score')
            if min_score is not None:
                queryset = queryset.filter(match_score__gte=min_score)
            if max_score is not None:
                queryset = queryset.filter(match_score__lte=max_score)
            
            # Status
            status = filter_form.cleaned_data.get('status')
            if status:
                queryset = queryset.filter(status=status)
            
            # Shortlisted
            if filter_form.cleaned_data.get('is_shortlisted'):
                queryset = queryset.filter(is_shortlisted=True)
            
            # Search
            search = filter_form.cleaned_data.get('search_query')
            if search:
                queryset = queryset.filter(
                    Q(resume__user__email__icontains=search) |
                    Q(resume__user__personal_profile__full_name__icontains=search)
                )
            
            # Sort (default to match score if no valid choice selected)
            sort_by = filter_form.cleaned_data.get('sort_by') or '-match_score'
            queryset = queryset.order_by(sort_by)
        else:
            queryset = queryset.order_by('-match_score')
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['session'] = self.get_session()
        context['filter_form'] = ScreeningResultFilterForm(self.request.GET)
        
        # Statistics
        queryset = self.get_queryset()
        context['stats'] = {
            'total': queryset.count(),
            'excellent': queryset.filter(match_score__gte=90).count(),
            'strong': queryset.filter(match_score__gte=80, match_score__lt=90).count(),
            'good': queryset.filter(match_score__gte=70, match_score__lt=80).count(),
            'shortlisted': queryset.filter(is_shortlisted=True).count(),
        }
        
        return context


class ScreeningResultDetailView(LoginRequiredMixin, CompanyOnlyMixin, DetailView):
    """View detailed screening result for a candidate."""
    
    model = ScreeningResult
    template_name = 'screening/result_detail.html'
    context_object_name = 'result'
    pk_url_kwarg = 'result_id'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        
        # Verify ownership
        if obj.session.company.user != self.request.user:
            raise PermissionDenied("You can only view your own screening results.")
        
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        result = self.object
        criteria = getattr(result.session, 'criteria', None)
        
        # Add note form
        context['note_form'] = ScreeningResultNoteForm()
        context['screening_answers'] = _normalize_screening_answers(result.screening_answers)
        
        # Parse match details
        context['match_details'] = result.match_details
        details = result.match_details or {}
        screening_percent = _analysis_percent(details, 'screening_answers_analysis')
        assessments_percent = _analysis_percent(details, 'assessments_analysis')
        experience_percent = _safe_percent(_resolve_nested_value(details, 'experience_match', 0))
        education_percent = _safe_percent(_resolve_nested_value(details, 'education_match', 0))
        semantic_percent = _safe_percent(_resolve_nested_value(details, 'semantic_similarity', 0))
        context['match_overview'] = {
            'experience': experience_percent,
            'education': education_percent,
            'semantic': semantic_percent
        }
        context['match_metrics'] = [
            {
                'label': 'Skills',
                'value': _safe_percent(_resolve_nested_value(details, 'skills_match.score', 0), already_percent=True),
                'color': 'bg-green-500'
            },
            {
                'label': 'Experience',
                'value': _safe_percent(_resolve_nested_value(details, 'experience_match', 0)),
                'color': 'bg-blue-500'
            },
            {
                'label': 'Education',
                'value': _safe_percent(_resolve_nested_value(details, 'education_match', 0)),
                'color': 'bg-purple-500'
            },
            {
                'label': 'Screening Answers',
                'value': screening_percent,
                'color': 'bg-indigo-500'
            },
            {
                'label': 'Assessments',
                'value': assessments_percent,
                'color': 'bg-emerald-500'
            }
        ]
        context['match_weights'] = {
            'skills': _criteria_weight(criteria, 'weight_skills', 0.4) * 100,
            'experience': _criteria_weight(criteria, 'weight_experience', 0.2) * 100,
            'education': _criteria_weight(criteria, 'weight_education', 0.2) * 100,
            'screening': _criteria_weight(criteria, 'weight_screening_questions', 0.1) * 100,
            'assessments': _criteria_weight(criteria, 'weight_assessments', 0.1) * 100
        }
        
        # Get resume data if available
        if result.resume:
            context['resume'] = result.resume
        
        # Interview questions (if Mistral AI is available)
        if result.match_score >= 70:  # Only for good matches
            try:
                from .mistral_client import mistral_client
                if mistral_client.is_available():
                    required_skills = criteria.required_skills if criteria else []
                    
                    context['interview_questions'] = mistral_client.generate_interview_questions(
                        job_title=result.job.title if result.job else "General Position",
                        required_skills=required_skills,
                        experience_level="senior",  # You can make this dynamic
                        num_questions=5
                    )
            except Exception as e:
                logger.error(f"Error generating interview questions: {e}")
        
        return context


class ScreeningSessionUpdateView(LoginRequiredMixin, CompanyOnlyMixin, ScreeningOwnerMixin, UpdateView):
    """Update screening session details."""
    
    model = ScreeningSession
    form_class = ScreeningSessionUpdateForm
    template_name = 'screening/session_update.html'
    context_object_name = 'session'

    @transaction.atomic
    def form_valid(self, form):
        """Handle valid update."""
        try:
            response = super().form_valid(form)
            
            logger.info(f"Session {self.object.id} updated")
            messages.success(self.request, "Session updated successfully.")
            
            return response
        except Exception as e:
            logger.error(f"Error updating session: {str(e)}", exc_info=True)
            messages.error(self.request, "An error occurred while updating.")
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('screening:session_detail', kwargs={'session_id': self.object.pk})


class ScreeningCriteriaUpdateView(LoginRequiredMixin, CompanyOnlyMixin, UpdateView):
    """Update screening criteria."""
    
    model = ScreeningCriteria
    form_class = ScreeningCriteriaForm
    template_name = 'screening/criteria_update.html'
    context_object_name = 'criteria'

    def get_object(self, queryset=None):
        session_pk = self.kwargs.get('pk')
        session = get_object_or_404(
            ScreeningSession,
            pk=session_pk,
            company__user=self.request.user
        )
        return get_object_or_404(ScreeningCriteria, session=session)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['session'] = self.object.session
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        """Handle valid update."""
        try:
            response = super().form_valid(form)
            
            logger.info(f"Criteria updated for session {self.object.session.id}")
            messages.success(self.request, "Criteria updated successfully.")
            
            return response
        except Exception as e:
            logger.error(f"Error updating criteria: {str(e)}", exc_info=True)
            messages.error(self.request, "An error occurred while updating.")
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('screening:session_detail', kwargs={'session_id': self.object.session.pk})


class ScreeningResultExportView(LoginRequiredMixin, CompanyOnlyMixin, FormView):
    """Export screening results."""
    
    form_class = ScreeningResultExportForm
    template_name = 'screening/export.html'

    def get_session(self):
        """Get the screening session."""
        pk = self.kwargs.get('pk')
        return get_object_or_404(
            ScreeningSession,
            pk=pk,
            company__user=self.request.user
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['session'] = self.get_session()
        return context

    def form_valid(self, form):
        """Generate export file."""
        session = self.get_session()
        export_format = form.cleaned_data['export_format']
        
        # Get filtered results
        results = session.results.select_related('resume', 'resume__user').all()
        
        # Apply filters
        if form.cleaned_data.get('only_shortlisted'):
            results = results.filter(is_shortlisted=True)
        
        min_score = form.cleaned_data.get('min_score')
        if min_score:
            results = results.filter(match_score__gte=min_score)
        
        # Export based on format
        if export_format == 'csv':
            return self._export_csv(session, results, form.cleaned_data)
        elif export_format == 'excel':
            return self._export_excel(session, results, form.cleaned_data)
        elif export_format == 'pdf':
            return self._export_pdf(session, results, form.cleaned_data)
        
        messages.error(self.request, "Invalid export format.")
        return self.form_invalid(form)

    def _export_csv(self, session, results, options):
        """Export to CSV."""
        import csv
        from io import StringIO
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="screening_{session.id}_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        headers = ['Rank', 'Candidate Name', 'Email', 'Match Score']
        if 'skills' in options['include_fields']:
            headers.append('Skills')
        if 'experience' in options['include_fields']:
            headers.append('Experience (Years)')
        if 'education' in options['include_fields']:
            headers.append('Education')
        headers.extend(['Shortlisted', 'Status'])
        
        writer.writerow(headers)
        
        # Write data
        for idx, result in enumerate(results.order_by('-match_score'), 1):
            row = [
                idx,
                result.resume.user.personal_profile.full_name if result.resume and hasattr(result.resume.user, 'personal_profile') else 'Unknown',
                result.resume.user.email if result.resume else 'Unknown',
                f"{result.match_score}%",
            ]
            
            if 'skills' in options['include_fields']:
                skills = result.match_details.get('resume_skills', [])
                row.append(', '.join(skills) if skills else 'N/A')
            
            if 'experience' in options['include_fields']:
                exp = result.match_details.get('resume_experience_years')
                row.append(f"{exp}" if exp else 'N/A')
            
            if 'education' in options['include_fields']:
                edu = result.match_details.get('resume_education')
                row.append(edu if edu else 'N/A')
            
            row.extend([
                'Yes' if result.is_shortlisted else 'No',
                result.get_status_display()
            ])
            
            writer.writerow(row)
        
        logger.info(f"CSV export for session {session.id}")
        return response

    def _export_excel(self, session, results, options):
        """Export to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Screening Results"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Write header
        headers = ['Rank', 'Candidate Name', 'Email', 'Match Score']
        if 'skills' in options['include_fields']:
            headers.append('Skills')
        if 'experience' in options['include_fields']:
            headers.append('Experience')
        if 'education' in options['include_fields']:
            headers.append('Education')
        headers.extend(['Shortlisted', 'Status'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for idx, result in enumerate(results.order_by('-match_score'), 1):
            row = idx + 1
            col = 1
            
            ws.cell(row=row, column=col, value=idx)
            col += 1
            
            ws.cell(row=row, column=col, value=result.resume.user.personal_profile.full_name if result.resume and hasattr(result.resume.user, 'personal_profile') else 'Unknown')
            col += 1
            
            ws.cell(row=row, column=col, value=result.resume.user.email if result.resume else 'Unknown')
            col += 1
            
            # Match score with color coding
            score_cell = ws.cell(row=row, column=col, value=result.match_score)
            if result.match_score >= 90:
                score_cell.fill = PatternFill(start_color="34C759", end_color="34C759", fill_type="solid")
            elif result.match_score >= 80:
                score_cell.fill = PatternFill(start_color="5AC8FA", end_color="5AC8FA", fill_type="solid")
            elif result.match_score >= 70:
                score_cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
            col += 1
            
            if 'skills' in options['include_fields']:
                skills = result.match_details.get('resume_skills', [])
                ws.cell(row=row, column=col, value=', '.join(skills) if skills else 'N/A')
                col += 1
            
            if 'experience' in options['include_fields']:
                exp = result.match_details.get('resume_experience_years')
                ws.cell(row=row, column=col, value=f"{exp}" if exp else 'N/A')
                col += 1
            
            if 'education' in options['include_fields']:
                edu = result.match_details.get('resume_education')
                ws.cell(row=row, column=col, value=edu if edu else 'N/A')
                col += 1
            
            ws.cell(row=row, column=col, value='Yes' if result.is_shortlisted else 'No')
            col += 1
            
            ws.cell(row=row, column=col, value=result.get_status_display())
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="screening_{session.id}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        logger.info(f"Excel export for session {session.id}")
        return response

    def _export_pdf(self, session, results, options):
        """Export to PDF using ReportLab."""
        try:
            from .pdf_export import ScreeningPDFExporter
            
            # Create PDF exporter
            exporter = ScreeningPDFExporter(session, results=results)
            pdf_buffer = exporter.generate()
            
            # Create response
            response = HttpResponse(pdf_buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="screening_{session.id}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            
            logger.info(f"PDF export for session {session.id} with {results.count()} results")
            return response
        
        except Exception as e:
            logger.error(f"PDF export failed: {str(e)}", exc_info=True)
            messages.error(self.request, f"PDF export failed: {str(e)}")
            return redirect('screening:results', session_id=session.pk)


# ===========================
# AJAX Views
# ===========================

class ScreeningProgressView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """AJAX endpoint for real-time progress updates."""

    def get(self, request, *args, **kwargs):
        """Get current progress."""
        try:
            pk = kwargs.get('pk')
            session = get_object_or_404(
                ScreeningSession,
                pk=pk,
                company__user=request.user
            )
            
            return JsonResponse({
                'success': True,
                'status': session.status,
                'progress': session.progress_percentage,
                'processed': session.processed_resumes,
                'total': session.total_resumes,
                'failed': session.failed_resumes,
                'average_score': session.average_match_score or 0,
            })
        except Exception as e:
            logger.error(f"Error getting progress: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)


class ScreeningResultShortlistToggleView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """AJAX endpoint for toggling shortlist status."""

    def post(self, request, *args, **kwargs):
        """Toggle shortlist status."""
        try:
            pk = kwargs.get('pk')
            logger.info(f"Shortlist toggle request: pk={pk}, user={request.user.id}")
            
            result = get_object_or_404(
                ScreeningResult,
                pk=pk,
                session__company__user=request.user
            )
            
            logger.info(f"Found result {result.id}, current shortlist status: {result.is_shortlisted}")
            result.toggle_shortlist()
            logger.info(f"Result {result.id} shortlist toggled to {result.is_shortlisted}")
            
            return JsonResponse({
                'success': True,
                'is_shortlisted': result.is_shortlisted
            })
        except ScreeningResult.DoesNotExist:
            logger.error(f"Result not found: pk={pk}")
            return JsonResponse({
                'success': False,
                'message': 'Result not found'
            }, status=404)
        except Exception as e:
            logger.error(f"Error toggling shortlist: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)


class BulkShortlistView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """AJAX endpoint for bulk shortlisting results."""

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        """Bulk shortlist/remove results."""
        try:
            import json
            data = json.loads(request.body)
            result_ids = data.get('result_ids', [])
            shortlist = data.get('shortlist', True)
            
            # Get all results for this company
            results = ScreeningResult.objects.filter(
                id__in=result_ids,
                session__company__user=request.user
            )
            
            # Update all matching results
            updated_count = 0
            for result in results:
                if result.is_shortlisted != shortlist:
                    result.is_shortlisted = shortlist
                    result.save()
                    updated_count += 1
            
            logger.info(f"Bulk shortlist operation: {updated_count} results updated to {shortlist}")
            
            return JsonResponse({
                'success': True,
                'updated': updated_count,
                'message': f'{updated_count} candidate(s) updated successfully'
            })
        except json.JSONDecodeError:
            logger.error("Invalid JSON in bulk shortlist request")
            return JsonResponse({
                'success': False,
                'message': 'Invalid request format'
            }, status=400)
        except Exception as e:
            logger.error(f"Error in bulk shortlist: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)


class ScreeningResultNoteAddView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """AJAX endpoint for adding notes to results."""

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        """Add note to result."""
        try:
            pk = kwargs.get('pk')
            result = get_object_or_404(
                ScreeningResult,
                pk=pk,
                session__company__user=request.user
            )
            
            note_text = request.POST.get('note')
            if not note_text:
                return JsonResponse({
                    'success': False,
                    'message': 'Note text is required'
                }, status=400)
            
            # Add note to result
            current_notes = result.notes or ''
            timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
            new_note = f"[{timestamp}] {note_text}"
            
            if current_notes:
                result.notes = f"{current_notes}\n{new_note}"
            else:
                result.notes = new_note
            
            result.save()
            
            logger.info(f"Note added to result {result.id}")
            
            return JsonResponse({
                'success': True,
                'note': new_note
            })
        except Exception as e:
            logger.error(f"Error adding note: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
from django.shortcuts import get_object_or_404
from django.views.generic import DetailView

class ScreeningAnalyticsView(LoginRequiredMixin, CompanyOnlyMixin, ScreeningOwnerMixin, DetailView):
    model = ScreeningSession
    template_name = 'screening/analytics.html'
    context_object_name = 'screening_session'
    pk_url_kwarg = 'session_id'

    def get_queryset(self):
        qs = super().get_queryset()

        try:
            company_profile = self.request.user.company_profile
        except CompanyProfile.DoesNotExist:
            logger.warning("Company profile required to view screening analytics.")
            raise PermissionDenied("Company profile required to access screening analytics.")

        return qs.filter(company=company_profile)

    def get_object(self, queryset=None):
        # This ensures we are loading the ScreeningSession by URL pk, not something else.
        queryset = queryset or self.get_queryset()
        return get_object_or_404(queryset, pk=self.kwargs.get(self.pk_url_kwarg))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        session = self.object  # guaranteed ScreeningSession instance

        context['screening_pk'] = str(session.pk)
        context['screening'] = session  # optional alias

        results = session.results.filter(
            status=ScreeningResultStatus.COMPLETED
        ).select_related('resume', 'resume__user')

        context['total_processed'] = results.count()
        context['total_failed'] = session.results.filter(status=ScreeningResultStatus.FAILED).count()
        context['total_pending'] = session.results.filter(status=ScreeningResultStatus.PENDING).count()
        context['average_match_score'] = session.average_match_score or 0

        context['high_matches'] = results.filter(match_score__gte=80).count()
        context['shortlisted_count'] = session.results.filter(is_shortlisted=True).count()

        context['score_distribution'] = {
            '0-59': results.filter(match_score__lt=60).count(),
            '60-69': results.filter(match_score__gte=60, match_score__lt=70).count(),
            '70-79': results.filter(match_score__gte=70, match_score__lt=80).count(),
            '80-89': results.filter(match_score__gte=80, match_score__lt=90).count(),
            '90-100': results.filter(match_score__gte=90).count(),
        }

        context['pie_data'] = {
            'completed': context['total_processed'],
            'failed': context['total_failed'],
            'pending': context['total_pending'],
        }

        skill_gaps = {}
        for r in results:
            for skill in (r.skills_gaps or []):
                skill_gaps[skill] = skill_gaps.get(skill, 0) + 1
        context['skill_gaps'] = dict(sorted(skill_gaps.items(), key=lambda x: x[1], reverse=True)[:5])

        matched_skills = {}
        for r in results:
            matched = (r.match_details or {}).get('skills_match', {}).get('matched', [])
            for skill in (matched or []):
                matched_skills[skill] = matched_skills.get(skill, 0) + 1
        context['top_matched_skills'] = dict(sorted(matched_skills.items(), key=lambda x: x[1], reverse=True)[:5])

        if session.completed_at and session.created_at:
            processing_time = session.completed_at - session.created_at
            context['processing_time'] = str(processing_time).split('.')[0]
            if context['total_processed'] > 0:
                avg_seconds = processing_time.total_seconds() / context['total_processed']
                context['avg_time_per_resume'] = f"{avg_seconds:.1f}s"

        experience_ranges = {'0-2': 0, '2-4': 0, '4-6': 0, '6-8': 0, '8+': 0}
        for r in results:
            exp = (r.match_details or {}).get('experience_match', {})
            years = exp.get('years') if isinstance(exp, dict) else None
            if years is None:
                continue
            if years < 2:
                experience_ranges['0-2'] += 1
            elif years < 4:
                experience_ranges['2-4'] += 1
            elif years < 6:
                experience_ranges['4-6'] += 1
            elif years < 8:
                experience_ranges['6-8'] += 1
            else:
                experience_ranges['8+'] += 1
        context['experience_distribution'] = experience_ranges

        return context


# ===========================
# Pipeline Integration Views
# ===========================

class PushToPipelineView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Push candidates to hiring pipeline."""

    def post(self, request, *args, **kwargs):
        """Handle pipeline push request."""
        try:
            pk = kwargs.get('pk')
            session = get_object_or_404(
                ScreeningSession,
                pk=pk,
                company__user=request.user
            )
            
            # Parse request data
            import json
            data = json.loads(request.body) if request.content_type == 'application/json' else request.POST
            result_ids = data.get('result_ids', [])
            job_id = data.get('job_id')
            
            if not result_ids or not job_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Missing result IDs or job ID'
                }, status=400)
            
            job = get_object_or_404(Job, id=job_id, company=session.company)
            
            # Create pipeline integrations
            pushed_count = 0
            error_count = 0
            
            for result_id in result_ids:
                try:
                    result = ScreeningResult.objects.get(
                        id=result_id,
                        session=session
                    )
                    
                    # Create or update pipeline integration
                    pipeline_int, created = PipelineIntegration.objects.update_or_create(
                        result=result,
                        defaults={
                            'job': job,
                            'company': session.company,
                            'status': PipelineStatus.PUSHED,
                        }
                    )
                    
                    # Log the push
                    logger.info(
                        f"Pushed {result.resume.user.email if result.resume else 'Unknown'} "
                        f"to pipeline for job {job.title}"
                    )
                    pushed_count += 1
                
                except ScreeningResult.DoesNotExist:
                    error_count += 1
                except Exception as e:
                    logger.error(f"Error pushing to pipeline: {str(e)}")
                    error_count += 1
            
            message = f"{pushed_count} candidate(s) pushed to pipeline"
            if error_count > 0:
                message += f" ({error_count} failed)"
            
            return JsonResponse({
                'success': error_count == 0,
                'pushed': pushed_count,
                'failed': error_count,
                'message': message
            })
        
        except ScreeningSession.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Screening session not found'
            }, status=404)
        
        except Exception as e:
            logger.error(f"Pipeline push error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


class BulkPushToPipelineView(LoginRequiredMixin, CompanyOnlyMixin, FormView):
    """Bulk push candidates to pipeline with strategy."""
    
    form_class = BulkPushToPipelineForm
    template_name = 'screening/bulk_push_pipeline.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        company = get_object_or_404(
            CompanyProfile,
            user=self.request.user
        )
        kwargs['company'] = company
        return kwargs

    def get_session(self):
        """Get screening session."""
        pk = self.kwargs.get('pk')
        return get_object_or_404(
            ScreeningSession,
            pk=pk,
            company__user=self.request.user
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['session'] = self.get_session()
        
        # Get selected results
        result_ids = self.request.GET.getlist('result_ids')
        if result_ids:
            context['results'] = ScreeningResult.objects.filter(
                id__in=result_ids,
                session=context['session']
            ).select_related('resume', 'resume__user')
        
        context['screening_session'] = context['session']
        return context

    def form_valid(self, form):
        """Process pipeline push."""
        session = self.get_session()
        
        # Parse result IDs
        result_ids_str = form.cleaned_data.get('result_ids', '')
        result_ids = [r.strip() for r in result_ids_str.split(',') if r.strip()]
        
        jobs = form.cleaned_data.get('jobs', [])
        strategy = form.cleaned_data.get('strategy', 'best_match')
        notify = form.cleaned_data.get('notify_recruiters', False)
        
        if not result_ids or not jobs:
            messages.error(self.request, 'No candidates or jobs selected')
            return self.form_invalid(form)
        
        # Get results
        results = ScreeningResult.objects.filter(
            id__in=result_ids,
            session=session
        )
        
        total_pushed = 0
        failed = 0
        
        # Process based on strategy
        if strategy == 'best_match':
            # Push each candidate to their best matching job
            for result in results:
                try:
                    best_job = self._find_best_job(result, jobs)
                    if best_job:
                        self._create_pipeline_integration(result, best_job)
                        total_pushed += 1
                except Exception as e:
                    logger.error(f"Error pushing result: {str(e)}")
                    failed += 1
        
        elif strategy == 'all':
            # Push candidate to all selected jobs
            for result in results:
                for job in jobs:
                    try:
                        self._create_pipeline_integration(result, job)
                        total_pushed += 1
                    except Exception as e:
                        logger.error(f"Error: {str(e)}")
                        failed += 1
        
        elif strategy == 'filtered':
            # Only push if match score meets threshold (70%)
            for result in results:
                if result.match_score >= 70:
                    best_job = self._find_best_job(result, jobs)
                    if best_job:
                        self._create_pipeline_integration(result, best_job)
                        total_pushed += 1
        
        # Create success message
        message = f'Successfully pushed {total_pushed} candidate(s) to pipeline'
        if failed > 0:
            message += f' ({failed} failed)'
        
        messages.success(self.request, message)
        
        return redirect('screening:results', session_id=session.pk)

    def _find_best_job(self, result, jobs):
        """Find best matching job for result."""
        # Could use semantic similarity in future
        # For now, return first job in list
        return jobs.first() if jobs.exists() else None

    def _create_pipeline_integration(self, result, job):
        """Create pipeline integration record."""
        pipeline_int, created = PipelineIntegration.objects.update_or_create(
            result=result,
            defaults={
                'job': job,
                'company': result.session.company,
                'status': PipelineStatus.PUSHED,
            }
        )
        return pipeline_int


class PipelineStatusUpdateView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Update pipeline status for a candidate."""

    def post(self, request, *args, **kwargs):
        """Handle status update."""
        try:
            import json
            data = json.loads(request.body)
            
            result_id = data.get('result_id')
            new_status = data.get('status')
            notes = data.get('notes', '')
            
            if not result_id or not new_status:
                return JsonResponse({
                    'success': False,
                    'message': 'Missing required fields'
                }, status=400)
            
            result = ScreeningResult.objects.select_related(
                'session'
            ).get(id=result_id, session__company__user=request.user)
            
            # Update pipeline integration
            pipeline_int = result.pipeline_integration
            
            if new_status == 'hired':
                pipeline_int.mark_as_hired()
            elif new_status == 'rejected':
                pipeline_int.mark_as_rejected()
            else:
                pipeline_int.status = new_status
                pipeline_int.save()
            
            if notes:
                pipeline_int.notes = notes
                pipeline_int.save(update_fields=['notes'])
            
            logger.info(f"Updated pipeline status for {result} to {new_status}")
            
            return JsonResponse({
                'success': True,
                'message': f'Status updated to {new_status}',
                'status': pipeline_int.get_status_display()
            })
        
        except ScreeningResult.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Result not found'
            }, status=404)
        
        except Exception as e:
            logger.error(f"Status update error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


# ===========================
# Real-Time Progress Views
# ===========================

class ProgressUpdateView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Polling endpoint for real-time progress updates."""

    def get(self, request, *args, **kwargs):
        """Get latest progress updates for session."""
        try:
            pk = kwargs.get('pk')
            session = get_object_or_404(
                ScreeningSession,
                pk=pk,
                company__user=request.user
            )

            # Get last update timestamp if provided
            since = request.GET.get('since', '')
            
            # Query progress updates
            updates_query = session.progress_updates.all()
            
            if since:
                try:
                    since_time = timezone.datetime.fromisoformat(since)
                    updates_query = updates_query.filter(created_at__gt=since_time)
                except (ValueError, AttributeError):
                    pass
            
            # Get recent updates (limit to last 50)
            updates = updates_query.order_by('-created_at')[:50]
            
            # Serialize updates
            updates_data = []
            for update in updates:
                updates_data.append({
                    'id': str(update.id),
                    'type': update.update_type,
                    'title': update.title,
                    'message': update.message,
                    'status': update.status,
                    'progress': update.progress_percent,
                    'current': update.current_item,
                    'total': update.total_items,
                    'error': update.error_message,
                    'timestamp': update.created_at.isoformat(),
                    'metadata': update.metadata,
                })
            
            # Get session status
            session_status = {
                'total_resumes': session.total_resumes,
                'processed_resumes': session.processed_resumes,
                'results_count': session.results.filter(status='completed').count(),
                'session_status': session.get_status_display(),
            }

            return JsonResponse({
                'success': True,
                'updates': updates_data,
                'session': session_status,
                'timestamp': timezone.now().isoformat(),
            })

        except Exception as e:
            logger.error(f"Progress update error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


class SessionStatsView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Get current session statistics for dashboard."""

    def get(self, request, *args, **kwargs):
        """Get session stats."""
        try:
            pk = kwargs.get('pk')
            session = get_object_or_404(
                ScreeningSession,
                pk=pk,
                company__user=request.user
            )

            # Calculate statistics
            results = session.results.all()
            completed = results.filter(status='completed')
            shortlisted = results.filter(is_shortlisted=True)
            
            avg_score = completed.aggregate(Avg('match_score'))['match_score__avg'] or 0
            high_matches = completed.filter(match_score__gte=80).count()
            
            # Pipeline stats
            pipeline_stats = {
                'pushed': session.results.filter(
                    pipeline_integration__status='pushed'
                ).count(),
                'hired': session.results.filter(
                    pipeline_integration__status='hired'
                ).count(),
                'rejected': session.results.filter(
                    pipeline_integration__status='rejected'
                ).count(),
            }

            return JsonResponse({
                'success': True,
                'stats': {
                    'total': results.count(),
                    'processed': completed.count(),
                    'shortlisted': shortlisted.count(),
                    'average_score': round(avg_score, 1),
                    'high_matches': high_matches,
                    'pending': results.filter(status='pending').count(),
                    'processing': results.filter(status='processing').count(),
                    'failed': results.filter(status='failed').count(),
                    'pipeline': pipeline_stats,
                    'session_status': session.get_status_display(),
                    'progress_percent': int((completed.count() / max(results.count(), 1)) * 100),
                }
            })

        except Exception as e:
            logger.error(f"Stats fetch error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


class CreateProgressUpdateView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Create progress update (for backend tasks)."""

    def post(self, request, *args, **kwargs):
        """Create a progress update."""
        try:
            import json
            data = json.loads(request.body)
            
            pk = kwargs.get('pk')
            session = get_object_or_404(
                ScreeningSession,
                pk=pk,
                company__user=request.user
            )

            update_type = data.get('update_type')
            title = data.get('title')
            message = data.get('message', '')
            progress = data.get('progress', 0)
            current = data.get('current_item')
            total = data.get('total_items')
            status = data.get('status', 'running')
            error = data.get('error_message', '')
            metadata = data.get('metadata', {})
            
            result_id = data.get('result_id')
            result = None
            if result_id:
                result = ScreeningResult.objects.filter(
                    id=result_id,
                    session=session
                ).first()

            # Create update
            progress_update = ProgressUpdate.objects.create(
                session=session,
                result=result,
                update_type=update_type,
                title=title,
                message=message,
                progress_percent=progress,
                current_item=current,
                total_items=total,
                status=status,
                error_message=error,
                metadata=metadata,
            )

            return JsonResponse({
                'success': True,
                'id': str(progress_update.id),
                'message': 'Progress update created'
            })

        except ScreeningSession.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Session not found'
            }, status=404)

        except Exception as e:
            logger.error(f"Create update error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


# ===========================
# AI Insight Views
# ===========================

class GenerateInsightView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Generate AI insights for a screening result."""

    def post(self, request, *args, **kwargs):
        """Generate an AI insight."""
        try:
            import json
            from .mistral_service import MistralAIService

            data = json.loads(request.body)
            result_id = data.get('result_id')
            insight_type = data.get('insight_type')

            # Get result
            result = get_object_or_404(
                ScreeningResult,
                id=result_id,
                session__company__user=request.user
            )

            # Check if insight already exists
            try:
                existing = AIInsight.objects.get(result=result, insight_type=insight_type)
                return JsonResponse({
                    'success': False,
                    'message': 'Insight already exists for this result and type'
                }, status=400)
            except AIInsight.DoesNotExist:
                pass

            # Initialize Mistral service
            service = MistralAIService()

            # Get resume text
            resume = result.resume
            if not resume:
                return JsonResponse({
                    'success': False,
                    'message': 'Resume not found for this result'
                }, status=400)

            resume_text = resume.extracted_text or resume.original_text or ""

            # Get job description
            job = result.session.job
            job_description = job.description if job else ""

            # Generate insight based on type
            if insight_type == 'interview_questions':
                ai_result = service.generate_interview_questions(resume_text, job_description)
                title = "Interview Questions"
                content_key = 'questions'

            elif insight_type == 'ai_notes':
                criteria = ScreeningCriteria.objects.filter(session=result.session).first()
                key_findings = criteria.custom_keywords if criteria else ""
                ai_result = service.generate_ai_notes(
                    resume_text,
                    result.match_score,
                    key_findings
                )
                title = "AI Notes"
                content_key = 'notes'

            elif insight_type == 'rejection_reasons':
                ai_result = service.generate_rejection_reasons(
                    resume_text,
                    job_description,
                    result.match_score
                )
                title = "Rejection Analysis"
                content_key = 'reasons'

            elif insight_type == 'resume_parsing':
                ai_result = service.parse_resume(resume_text)
                title = "Parsed Resume Data"
                content_key = 'parsed_data'

            else:
                return JsonResponse({
                    'success': False,
                    'message': f'Unknown insight type: {insight_type}'
                }, status=400)

            if not ai_result.get('success'):
                # Log error
                logger.error(f"AI insight generation failed: {ai_result.get('error')}")
                
                # Create progress update for failure
                ProgressUpdate.objects.create(
                    session=result.session,
                    result=result,
                    update_type='error_occurred',
                    title=f'AI {title} Generation Failed',
                    message=ai_result.get('error', 'Unknown error'),
                    status='failed',
                    error_message=ai_result.get('error', ''),
                )

                return JsonResponse({
                    'success': False,
                    'message': ai_result.get('error', 'Failed to generate insight')
                }, status=500)

            # Create AI insight
            insight_content = {
                content_key: ai_result.get(content_key, [])
            }

            # Add optional summary
            if 'summary' in ai_result:
                insight_content['summary'] = ai_result['summary']
            if 'recommendation' in ai_result:
                insight_content['recommendation'] = ai_result['recommendation']

            ai_insight = AIInsight.objects.create(
                result=result,
                insight_type=insight_type,
                title=title,
                content=insight_content,
                summary=ai_result.get('summary', ai_result.get('recommendation', '')),
                tokens_used=ai_result.get('tokens_used', 0),
                generation_time=ai_result.get('generation_time', 0),
                confidence_score=0.85,  # Default confidence
            )

            # Create progress update
            ProgressUpdate.objects.create(
                session=result.session,
                result=result,
                update_type='result_analyzed',
                title=f'AI {title} Generated',
                message=f'Successfully generated {title.lower()} for {result.candidate_name}',
                status='completed',
                progress_percent=100,
                metadata={
                    'insight_id': str(ai_insight.id),
                    'tokens_used': ai_result.get('tokens_used', 0),
                }
            )

            return JsonResponse({
                'success': True,
                'insight_id': str(ai_insight.id),
                'title': title,
                'content': insight_content,
                'tokens_used': ai_result.get('tokens_used', 0),
                'generation_time': ai_result.get('generation_time', 0),
                'message': f'{title} generated successfully'
            })

        except ScreeningResult.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Result not found'
            }, status=404)

        except Exception as e:
            logger.error(f"Insight generation error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


class BatchGenerateInsightView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Generate insights for multiple results."""

    def post(self, request, *args, **kwargs):
        """Generate insights for multiple results."""
        try:
            import json
            from .mistral_service import MistralAIService

            data = json.loads(request.body)
            result_ids = data.get('result_ids', [])
            insight_types = data.get('insight_types', ['interview_questions', 'ai_notes'])
            
            session_id = kwargs.get('pk')
            session = get_object_or_404(
                ScreeningSession,
                id=session_id,
                company__user=request.user
            )

            # Get results
            results = ScreeningResult.objects.filter(
                id__in=result_ids,
                session=session
            )

            if not results.exists():
                return JsonResponse({
                    'success': False,
                    'message': 'No results found'
                }, status=400)

            # Initialize service
            service = MistralAIService()

            insights_created = []
            errors = []

            for result in results:
                for insight_type in insight_types:
                    try:
                        # Skip if already exists
                        if AIInsight.objects.filter(result=result, insight_type=insight_type).exists():
                            continue

                        # Get resume
                        resume = result.resume
                        if not resume:
                            continue

                        resume_text = resume.extracted_text or resume.original_text or ""
                        job = session.job
                        job_description = job.description if job else ""

                        # Generate insight
                        if insight_type == 'interview_questions':
                            ai_result = service.generate_interview_questions(resume_text, job_description)
                            title = "Interview Questions"
                            content_key = 'questions'

                        elif insight_type == 'ai_notes':
                            criteria = ScreeningCriteria.objects.filter(session=session).first()
                            key_findings = criteria.custom_keywords if criteria else ""
                            ai_result = service.generate_ai_notes(resume_text, result.match_score, key_findings)
                            title = "AI Notes"
                            content_key = 'notes'

                        elif insight_type == 'rejection_reasons':
                            ai_result = service.generate_rejection_reasons(resume_text, job_description, result.match_score)
                            title = "Rejection Analysis"
                            content_key = 'reasons'

                        elif insight_type == 'resume_parsing':
                            ai_result = service.parse_resume(resume_text)
                            title = "Parsed Resume Data"
                            content_key = 'parsed_data'

                        else:
                            continue

                        if not ai_result.get('success'):
                            errors.append({
                                'result_id': str(result.id),
                                'type': insight_type,
                                'error': ai_result.get('error', 'Unknown error')
                            })
                            continue

                        # Create insight
                        insight_content = {content_key: ai_result.get(content_key, [])}
                        if 'summary' in ai_result:
                            insight_content['summary'] = ai_result['summary']
                        if 'recommendation' in ai_result:
                            insight_content['recommendation'] = ai_result['recommendation']

                        ai_insight = AIInsight.objects.create(
                            result=result,
                            insight_type=insight_type,
                            title=title,
                            content=insight_content,
                            summary=ai_result.get('summary', ''),
                            tokens_used=ai_result.get('tokens_used', 0),
                            generation_time=ai_result.get('generation_time', 0),
                            confidence_score=0.85,
                        )

                        insights_created.append({
                            'result_id': str(result.id),
                            'insight_id': str(ai_insight.id),
                            'type': insight_type,
                            'title': title,
                        })

                    except Exception as e:
                        logger.error(f"Error generating {insight_type} for result {result.id}: {str(e)}")
                        errors.append({
                            'result_id': str(result.id),
                            'type': insight_type,
                            'error': str(e)
                        })

            # Create progress update
            ProgressUpdate.objects.create(
                session=session,
                update_type='result_analyzed',
                title='Batch AI Insights Generated',
                message=f'Generated {len(insights_created)} insights for {len(results)} candidates',
                status='completed',
                progress_percent=100,
                current_item=len(results),
                total_items=len(results),
                metadata={
                    'insights_created': len(insights_created),
                    'errors': len(errors),
                }
            )

            return JsonResponse({
                'success': True,
                'insights_created': insights_created,
                'errors': errors,
                'total_created': len(insights_created),
                'total_errors': len(errors),
                'message': f'Generated {len(insights_created)} insights'
            })

        except Exception as e:
            logger.error(f"Batch insight generation error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


class RetrieveInsightView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Retrieve AI insights for a result."""

    def get(self, request, *args, **kwargs):
        """Get insights for a result."""
        try:
            result_id = kwargs.get('result_id')
            insight_type = request.GET.get('type')

            # Get result
            result = get_object_or_404(
                ScreeningResult,
                id=result_id,
                session__company__user=request.user
            )

            # Get insights
            if insight_type:
                insights = AIInsight.objects.filter(result=result, insight_type=insight_type)
            else:
                insights = AIInsight.objects.filter(result=result)

            insights_data = []
            for insight in insights:
                insights_data.append({
                    'id': str(insight.id),
                    'type': insight.insight_type,
                    'title': insight.title,
                    'content': insight.content,
                    'summary': insight.summary,
                    'confidence': insight.confidence_score,
                    'is_approved': insight.is_approved,
                    'is_used': insight.is_used,
                    'created_at': insight.created_at.isoformat(),
                })

            return JsonResponse({
                'success': True,
                'insights': insights_data,
                'count': len(insights_data),
            })

        except ScreeningResult.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Result not found'
            }, status=404)

        except Exception as e:
            logger.error(f"Retrieve insight error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


class ApproveInsightView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Approve an AI insight."""

    def post(self, request, *args, **kwargs):
        """Approve an insight."""
        try:
            import json
            data = json.loads(request.body)
            insight_id = data.get('insight_id')

            insight = get_object_or_404(
                AIInsight,
                id=insight_id,
                result__session__company__user=request.user
            )

            insight.mark_approved()

            return JsonResponse({
                'success': True,
                'message': 'Insight approved',
                'is_approved': insight.is_approved,
            })

        except AIInsight.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Insight not found'
            }, status=404)

        except Exception as e:
            logger.error(f"Approve insight error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


class FeedbackInsightView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """Submit feedback on an AI insight."""

    def post(self, request, *args, **kwargs):
        """Submit feedback."""
        try:
            import json
            data = json.loads(request.body)
            insight_id = data.get('insight_id')
            rating = data.get('rating')
            comment = data.get('comment', '')

            insight = get_object_or_404(
                AIInsight,
                id=insight_id,
                result__session__company__user=request.user
            )

            # Create feedback
            feedback = InsightFeedback.objects.create(
                insight=insight,
                rating=rating,
                comment=comment,
                user=request.user,
            )

            return JsonResponse({
                'success': True,
                'feedback_id': str(feedback.id),
                'message': 'Feedback recorded',
            })

        except AIInsight.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Insight not found'
            }, status=404)

        except Exception as e:
            logger.error(f"Feedback error: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


class SessionResultsAPIView(LoginRequiredMixin, CompanyOnlyMixin, View):
    """API endpoint to fetch current screening results with live match scores."""

    def get(self, request, *args, **kwargs):
        """Get all results for a session with current match scores."""
        try:
            session_id = kwargs.get('session_id')
            limit = request.GET.get('limit', 10)
            
            try:
                limit = int(limit)
            except (ValueError, TypeError):
                limit = 10
            
            # Get session
            session = get_object_or_404(
                ScreeningSession,
                pk=session_id,
                company__user=request.user
            )
            
            # Get results with current data from database
            results = session.results.select_related(
                'resume',
                'resume__user',
                'resume__user__personal_profile',
                'job',
                'application'
            ).order_by('-match_score')[:limit]
            
            # Serialize results
            results_data = []
            for result in results:
                candidate_name = result.candidate_name
                candidate_email = ''
                if result.resume:
                    candidate_email = result.resume.user.email
                elif result.application:
                    candidate_email = result.application.user.email
                
                results_data.append({
                    'id': str(result.id),
                    'candidate_name': candidate_name,
                    'candidate_email': candidate_email,
                    'match_score': result.match_score,
                    'status': result.status,
                    'is_shortlisted': result.is_shortlisted,
                    'skills_match': result.skills_match,
                    'experience_match': result.experience_match,
                    'education_match': result.education_match,
                    'processed_at': result.processed_at.isoformat() if result.processed_at else None,
                    'error_message': result.error_message,
                })
            
            return JsonResponse({
                'success': True,
                'results': results_data,
                'total': session.results.count(),
                'session_id': str(session.id),
                'session_status': session.status,
                'average_match_score': session.average_match_score,
                'processed_resumes': session.processed_resumes,
                'total_resumes': session.total_resumes,
            })
        
        except ScreeningSession.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Session not found'
            }, status=404)
        
        except Exception as e:
            logger.error(f"Error fetching session results: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)


        return context
